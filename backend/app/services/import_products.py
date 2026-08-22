"""Excel (.xlsx) / CSV ფაილიდან პროდუქტების წაკითხვა და ვალიდაცია.

parse_products_file(content, filename) -> (products, errors)
  products: list[dict] — ვალიდური პროდუქტები (shop_id-ის გარეშე)
  errors:   list[{row, message}] — მწკრივების შეცდომები
სტრუქტურულ პრობლემაზე (ფაილის ტიპი, აკლია სვეტი) აგდებს ValueError-ს.
"""
import csv
import io

from openpyxl import load_workbook

# სვეტების სახელების სინონიმები (ქართული + ინგლისური, case-insensitive)
COLUMN_ALIASES = {
    "name": {"სახელი", "დასახელება", "name", "title", "პროდუქტი"},
    "price": {"ფასი", "price", "ღირებულება"},
    "quantity": {"მარაგი", "რაოდენობა", "quantity", "qty", "stock"},
    "description": {"დეტალები", "აღწერა", "description", "details", "კომენტარი"},
    "sku": {"sku", "კოდი", "არტიკული"},
}
MAX_ROWS = 5000


def _norm(value) -> str:
    return ("" if value is None else str(value)).strip()


def _map_headers(headers) -> dict:
    mapping = {}
    for idx, h in enumerate(headers):
        hn = _norm(h).lower()
        for field, aliases in COLUMN_ALIASES.items():
            if hn in {a.lower() for a in aliases}:
                mapping[field] = idx
    return mapping


def _read_csv(content: bytes):
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValueError("CSV-ის წაკითხვა ვერ მოხერხდა — შეინახე UTF-8 კოდირებით.")
    rows = []
    for row in csv.reader(io.StringIO(text)):
        rows.append(row)
        if len(rows) > MAX_ROWS + 1:  # იხ. P2-11 კომენტარი _read_xlsx-ში
            raise ValueError(f"ძალიან ბევრი მწკრივი (მაქს. {MAX_ROWS}).")
    return rows


def _read_xlsx(content: bytes):
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ValueError("Excel ფაილის გახსნა ვერ მოხერხდა — დარწმუნდი, რომ .xlsx ფორმატია.")
    ws = wb.active
    if ws is None:
        raise ValueError("Excel ფაილში აქტიური ფურცელი ვერ მოიძებნა.")
    # ⚠️ რევიუ P2-11: ჭერი კითხვისასვე მოქმედებს. ადრე მთელი ფაილი ჯერ მეხსიერებაში
    # ჩაიტვირთებოდა და MAX_ROWS მხოლოდ ამის შემდეგ მოწმდებოდა — .xlsx zip-ია,
    # ანუ 5MB მილიონ მწკრივამდე იშლებოდა და instance-ს მეხსიერებას ამოწურავდა.
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if len(rows) > MAX_ROWS + 1:  # +1 = სათაურის მწკრივი
            raise ValueError(f"ძალიან ბევრი მწკრივი (მაქს. {MAX_ROWS}).")
    return rows


def read_rows(content: bytes, filename: str):
    """ფაილს კითხულობს და აბრუნებს არა-ცარიელ მწკრივებს (list[list])."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _read_csv(content)
    elif name.endswith(".xlsx"):
        rows = _read_xlsx(content)
    else:
        raise ValueError("მხარდაჭერილია მხოლოდ .csv და .xlsx ფაილები.")

    # სრულიად ცარიელი მწკრივების გამოტოვება
    rows = [r for r in rows if any(_norm(c) != "" for c in r)]
    if not rows:
        raise ValueError("ფაილი ცარიელია.")
    return rows


def preview_file(content: bytes, filename: str, sample: int = 5) -> dict:
    """ფაილის სათაურები + ავტო-ამოცნობილი mapping + პირველი მწკრივები.

    გამოიყენება „სვეტების მორგების" UI-სთვის: გამყიდველი ხედავს თავის სვეტებს
    და ადასტურებს/ასწორებს რომელია სახელი/ფასი/მარაგი.
    """
    rows = read_rows(content, filename)
    headers = [_norm(h) for h in rows[0]]
    detected = _map_headers(rows[0])  # {field: col_index}
    data_rows = rows[1:]
    preview = [[_norm(c) for c in r] for r in data_rows[:sample]]
    return {
        "headers": headers,
        "detected": detected,
        "preview": preview,
        "total_rows": len(data_rows),
    }


def parse_products_file(
    content: bytes,
    filename: str,
    mapping_override: dict | None = None,
    extra_cols: list | None = None,
):
    """ფაილიდან პროდუქტების წაკითხვა/ვალიდაცია.

    mapping_override: {field: col_index} — გამყიდვლის მიერ არჩეული სვეტები.
      თუ მითითებულია, ავტო-ამოცნობის ნაცვლად ეს გამოიყენება.
    extra_cols: [col_index, ...] — დამატებითი სვეტები, რომლებიც აღწერის ტექსტში
      ჩაიწერება „სათაური: მნიშვნელობა" ფორმატით (ბაზაში ახალი სვეტი არ ემატება).
    """
    rows = read_rows(content, filename)

    headers = rows[0]
    if mapping_override:
        mapping = {k: v for k, v in mapping_override.items() if v is not None}
    else:
        mapping = _map_headers(headers)
    if mapping.get("name") is None:
        found = ", ".join(_norm(h) for h in headers if _norm(h))
        raise ValueError(
            "აუცილებელია მიუთითო რომელი სვეტია 'სახელი'. ნაპოვნი სათაურები: "
            + (found or "(ცარიელი)")
        )

    data_rows = rows[1:]
    if len(data_rows) > MAX_ROWS:
        raise ValueError(f"ძალიან ბევრი მწკრივი (მაქს. {MAX_ROWS}).")

    products, errors = [], []
    seen_skus = set()

    # დამატებითი სვეტები აღწერისთვის — გამოვრიცხოთ ის, რაც უკვე mapping-შია
    used_cols = {v for v in mapping.values() if v is not None}
    clean_extra = [c for c in (extra_cols or []) if c is not None and c not in used_cols]

    def build_description(row):
        d_idx = mapping.get("description")
        base = _norm(row[d_idx]) if (d_idx is not None and d_idx < len(row)) else ""
        extras = []
        for idx in clean_extra:
            if idx >= len(row):
                continue
            val = _norm(row[idx])
            if not val:
                continue
            label = _norm(headers[idx]) if idx < len(headers) else ""
            extras.append(f"{label}: {val}" if label else val)
        extra_text = ", ".join(extras)
        if base and extra_text:
            return f"{base} — {extra_text}"
        return base or extra_text or None

    for offset, row in enumerate(data_rows):
        line = offset + 2  # Excel-ის ნუმერაცია (1 = სათაური)

        def cell(field):
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name_s = _norm(cell("name"))
        if not name_s:
            errors.append({"row": line, "message": "სახელი ცარიელია"})
            continue

        price = 0.0
        price_raw = _norm(cell("price"))
        if price_raw:
            try:
                price = float(price_raw.replace(",", ".").replace(" ", ""))
            except ValueError:
                errors.append({"row": line, "message": f"ფასი არ არის რიცხვი: „{price_raw}"})
                continue
            if price < 0:
                errors.append({"row": line, "message": "ფასი უარყოფითია"})
                continue

        qty = 0
        qty_raw = _norm(cell("quantity"))
        if qty_raw:
            try:
                qty = int(float(qty_raw.replace(",", ".")))
            except ValueError:
                errors.append({"row": line, "message": f"მარაგი არ არის რიცხვი: „{qty_raw}"})
                continue
            if qty < 0:
                errors.append({"row": line, "message": "მარაგი უარყოფითია"})
                continue

        sku = _norm(cell("sku")) or None
        if sku:
            if sku in seen_skus:
                errors.append({"row": line, "message": f"SKU მეორდება ფაილში: „{sku}"})
                continue
            seen_skus.add(sku)

        products.append({
            "name": name_s,
            "price": price,
            "quantity": qty,
            "description": build_description(row),
            "sku": sku,
            "is_active": True,
        })

    return products, errors
